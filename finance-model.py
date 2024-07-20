import streamlit as st
import numpy as np
import numpy_financial as npf
import pandas as pd
import plotly.graph_objects as go
import io
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side
import streamlit as st
import numpy as np
import numpy_financial as npf
import pandas as pd
import plotly.graph_objects as go
from openai import OpenAI
import io
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side
from datetime import datetime


ASSISTANT_ID = st.secrets["NY_ADVISOR"],
client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

# Initialize session state
if 'project_data' not in st.session_state:
    st.session_state['project_data'] = {}
if 'calculated_results' not in st.session_state:
    st.session_state['calculated_results'] = None

def calculate_ppa_price(project_data):
    capacity = project_data['capacity']
    capex = project_data['capex']
    opex = project_data['opex']
    lifespan = project_data['lifespan']
    degradation_rate = project_data['degradation_rate']
    annual_generation = project_data['annual_generation']
    target_irr = project_data['target_irr'] / 100  # Convert to decimal
    inverter_cost = project_data['inverter_cost']
    inverter_lifetime = project_data['inverter_lifetime']
    incentives = project_data['incentives']
    inflation_rate = project_data['inflation_rate']
    incentive_payback_period = project_data['incentive_payback_period']

    # Calculate total incentives
    itc_amount = capex * (incentives['itc_rate'] + incentives['additional_itc'])
    ny_incentives = calculate_ny_incentives(project_data) if incentives['ny_selected'] else 0
    total_incentives = itc_amount + ny_incentives

    def npv_function(ppa_price):
        cash_flows = []
        for year in range(lifespan):
            year_generation = annual_generation * (1 - degradation_rate) ** year
            revenue = year_generation * ppa_price
            annual_opex = opex * (1 + inflation_rate) ** year
            inverter_replacement = inverter_cost if (year + 1) % inverter_lifetime == 0 and year + 1 != lifespan else 0
            
            annual_incentive = total_incentives / incentive_payback_period if year < incentive_payback_period else 0
            
            cash_flow = revenue - annual_opex - inverter_replacement + annual_incentive
            cash_flows.append(cash_flow)
        
        cash_flows.insert(0, -capex)
        
        return npf.npv(target_irr, cash_flows)

    low, high = 0, 1
    while high - low > 0.0001:
        mid = (low + high) / 2
        if npv_function(mid) < 0:
            low = mid
        else:
            high = mid
    
    return (low + high) / 2


def calculate_ny_incentives(project_data):
    capacity = project_data['capacity']
    incentives = project_data['incentives']
    
    total_incentive = 0
    
    if incentives['ny_selected']:
        # ICSA
        if incentives['icsa_eligible']:
            if incentives['project_location'] == 'Upstate':
                total_incentive += capacity * 1000 * incentives['icsa_rate_upstate']
            else:  # Con Edison
                total_incentive += capacity * 1000 * incentives['icsa_rate_con_ed']
        
        # Landfill/Brownfield Adder
        if incentives['brownfield_eligible']:
            total_incentive += capacity * 1000 * 0.15
        
        # Prevailing Wage Adder
        if incentives['prevailing_wage_eligible']:
            if incentives['project_location'] == 'Upstate':
                total_incentive += capacity * 1000 * 0.125
            else:  # Con Edison
                total_incentive += capacity * 1000 * 0.20
        
        # NYSERDA Community Adder
        if incentives['nyserda_community_eligible']:
            total_incentive += capacity * 1000 * 0.07
    
    return total_incentive


def calculate_irr(project_data, ppa_price):
    capacity = project_data['capacity']
    capex = project_data['capex']
    opex = project_data['opex']
    lifespan = project_data['lifespan']
    degradation_rate = project_data['degradation_rate']
    annual_generation = project_data['annual_generation']
    inverter_cost = project_data['inverter_cost']
    inverter_lifetime = project_data['inverter_lifetime']
    incentives = project_data['incentives']
    inflation_rate = project_data['inflation_rate']

    # Calculate ITC amount
    itc_rate = incentives['itc_rate'] + incentives['additional_itc']
    itc_amount = capex * itc_rate

    # Calculate NY-specific incentives
    ny_incentives = calculate_ny_incentives(project_data)

    # Apply ITC and other incentives to reduce CapEx
    adjusted_capex = capex - itc_amount - ny_incentives

    cash_flows = [-adjusted_capex]
    for year in range(lifespan):
        year_generation = annual_generation * (1 - degradation_rate) ** year
        revenue = year_generation * ppa_price
        annual_opex = opex * (1 + inflation_rate) ** year
        inverter_replacement = inverter_cost if (year + 1) % inverter_lifetime == 0 and year + 1 != lifespan else 0
        cash_flow = revenue - annual_opex - inverter_replacement
        cash_flows.append(cash_flow)

    return npf.irr(cash_flows)




def generate_cash_flow_df(project_data, ppa_price):
    capacity = project_data['capacity']
    capex = project_data['capex']
    opex = project_data['opex']
    lifespan = project_data['lifespan']
    degradation_rate = project_data['degradation_rate']
    annual_generation = project_data['annual_generation']
    inverter_cost = project_data['inverter_cost']
    inverter_lifetime = project_data['inverter_lifetime']
    incentives = project_data['incentives']
    inflation_rate = project_data['inflation_rate']
    incentive_payback_period = project_data['incentive_payback_period']

    # Calculate total incentives
    itc_amount = capex * (incentives['itc_rate'] + incentives['additional_itc'])
    ny_incentives = calculate_ny_incentives(project_data) if incentives['ny_selected'] else 0
    total_incentives = itc_amount + ny_incentives
    annual_incentive = total_incentives / incentive_payback_period

    data = []
    cumulative_cash_flow = -capex

    for year in range(lifespan):
        year_generation = annual_generation * (1 - degradation_rate) ** year
        revenue = year_generation * ppa_price
        annual_opex = opex * (1 + inflation_rate) ** year
        inverter_replacement = inverter_cost if (year + 1) % inverter_lifetime == 0 and year + 1 != lifespan else 0
        
        incentive = annual_incentive if year < incentive_payback_period else 0
        
        cash_flow = revenue - annual_opex - inverter_replacement + incentive
        cumulative_cash_flow += cash_flow
        
        data.append({
            'Year': year + 1,
            'Annual Generation (kWh)': year_generation,
            'Revenue': revenue,
            'OpEx': annual_opex,
            'Inverter Replacement': inverter_replacement,
            'Incentive': incentive,
            'Cash Flow': cash_flow,
            'Cumulative Cash Flow': cumulative_cash_flow
        })

    return pd.DataFrame(data)

def plot_cash_flows(df):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['Year'], y=df['Cash Flow'], mode='lines+markers', name='Annual Cash Flow'))
    fig.add_trace(go.Scatter(x=df['Year'], y=df['Cumulative Cash Flow'], mode='lines+markers', name='Cumulative Cash Flow'))
    fig.update_layout(
        title='Project Cash Flows',
        xaxis_title='Year',
        yaxis_title='Cash Flow ($)',
        hovermode='x unified'
    )
    return fig

def export_to_excel(df, fig):
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cash Flows', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Cash Flows']

        # Format the worksheet
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
    
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Add styles
        header_font = Font(bold=True)
        centered_alignment = Alignment(horizontal='center')
        border_style = Border(left=Side(style='thin'), 
                              right=Side(style='thin'), 
                              top=Side(style='thin'), 
                              bottom=Side(style='thin'))

        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.border = border_style

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border_style

        # Add chart
        chart_sheet = workbook.create_sheet(title='Cash Flow Chart')
        chart = LineChart()
        chart.title = 'Project Cash Flows'
        chart.x_axis.title = 'Year'
        chart.y_axis.title = 'Cash Flow ($)'

        data = Reference(worksheet, min_col=6, min_row=1, max_col=7, max_row=len(df)+1)
        chart.add_data(data, titles_from_data=True)
        dates = Reference(worksheet, min_col=1, min_row=2, max_row=len(df)+1)
        chart.set_categories(dates)

        chart_sheet.add_chart(chart, 'B2')

    return buffer

def calculate_results(project_data):
    if project_data['use_custom_ppa']:
        ppa_price = project_data['custom_ppa_price']
        irr = calculate_irr(project_data, ppa_price)
    else:
        ppa_price = calculate_ppa_price(project_data)
        irr = project_data['target_irr']

    df = generate_cash_flow_df(project_data, ppa_price)
    
    total_generation = df['Annual Generation (kWh)'].sum()
    total_revenue = df['Revenue'].sum()
    total_opex = df['OpEx'].sum()
    total_incentives = df['Incentive'].sum()
    net_profit = total_revenue - total_opex - project_data['capex'] + total_incentives

    results = {
        'ppa_price': ppa_price,
        'irr': irr,
        'df': df,
        'total_generation': total_generation,
        'total_revenue': total_revenue,
        'total_opex': total_opex,
        'total_incentives': total_incentives,
        'net_profit': net_profit
    }
    
    return results

def display_results(results, project_data):
    if project_data['use_custom_ppa']:
        st.success(f'The IRR for the custom PPA price of ${results["ppa_price"]:.4f} per kWh is: {results["irr"]*100:.2f}%')
    else:
        st.success(f'The required PPA price to achieve a {project_data["target_irr"]}% IRR is: ${results["ppa_price"]:.4f} per kWh')

    st.subheader('Project Metrics')
    col1, col2 = st.columns(2)
    with col1:
        st.metric('Total Generation', f'{results["total_generation"]/1e6:.2f} GWh')
        st.metric('Total Revenue', f'${results["total_revenue"]/1e6:.2f}M')
    with col2:
        st.metric('Total OpEx', f'${results["total_opex"]/1e6:.2f}M')
        st.metric('Net Profit', f'${results["net_profit"]/1e6:.2f}M')

    st.metric('Total Incentives', f'${results["total_incentives"]:,.2f}')

    st.subheader('Cash Flow Table')
    st.dataframe(results['df'].style.format({
        'Annual Generation (kWh)': '{:,.0f}',
        'Revenue': '${:,.2f}',
        'OpEx': '${:,.2f}',
        'Inverter Replacement': '${:,.2f}',
        'Incentive': '${:,.2f}',
        'Cash Flow': '${:,.2f}',
        'Cumulative Cash Flow': '${:,.2f}'
    }))

    st.subheader('Interactive Cash Flow Graph')
    fig = plot_cash_flows(results['df'])
    st.plotly_chart(fig, use_container_width=True)

def main():
    st.title('PPA <> IRR Calculator ')

    st.sidebar.header('Project Inputs')
    with st.sidebar.expander("Project Specifications", expanded=True):
        project_capacity = st.number_input('Project Capacity (MW)', value=2.0, min_value=0.1, help="Total capacity of the solar project in megawatts")
        capex = st.number_input('Capital Expenditure (USD)', value=5000000, min_value=1, help="Total upfront cost for the project, including equipment and installation")
        opex = st.number_input('Annual Operational Expenditure (USD)', value=50000, min_value=0, help="Yearly cost for maintaining and operating the solar plant")
        project_lifespan = st.number_input('Project Lifespan (years)', value=25, min_value=1, max_value=50, help="Expected operational life of the project")

    with st.sidebar.expander("Performance Parameters", expanded=True):
        generation_method = st.radio("Generation Calculation Method", ["Annual Generation", "Peak Sun Hours"])
        
        if generation_method == "Annual Generation":
            annual_generation = st.number_input('Annual Generation (kWh)', value=3000000, min_value=1, help="Expected annual electricity generation")
        else:
            peak_sun_hours = st.number_input('Peak Sun Hours per Day', value=5.0, min_value=1.0, max_value=12.0, help="Average daily hours of peak sunlight")
            system_efficiency = st.number_input('System Efficiency (%)', value=80.0, min_value=10.0, max_value=100.0, help="Overall efficiency of the solar power system") / 100
            annual_generation = project_capacity * 1000 * peak_sun_hours * 365 * system_efficiency

        degradation_rate = st.number_input('Annual Degradation Rate (%)', value=0.5, min_value=0.0, max_value=5.0, help="Yearly reduction in solar panel efficiency") / 100

    with st.sidebar.expander("Financial Parameters", expanded=True):
        target_irr = st.slider('Target IRR (%)', min_value=5.0, max_value=15.0, value=9.5, step=0.1, help="Internal Rate of Return goal for the project")
        use_custom_ppa = st.checkbox("Use Custom PPA Price", value=False)
        if use_custom_ppa:
            custom_ppa_price = st.number_input('Custom PPA Price ($/kWh)', value=0.10, min_value=0.01, max_value=1.0, step=0.01, help="Enter a custom PPA price to calculate IRR")
        else:
            custom_ppa_price = None
        inverter_cost = st.number_input('Inverter Replacement Cost (USD)', value=500000, min_value=0, help="Cost to replace inverters")
        inverter_lifetime = st.number_input('Inverter Lifetime (years)', value=10, min_value=1, max_value=25, help="Expected operational life of inverters")
        inflation_rate = st.number_input('Inflation Rate (%)', value=2.0, min_value=0.0, max_value=10.0, step=0.1, help="Annual inflation rate for OpEx") / 100
        incentive_payback_period = st.number_input('Incentive Payback Period (years)', value=1, min_value=1, max_value=project_lifespan, help="Period over which incentives are applied")

    with st.sidebar.expander("Incentives", expanded=True):
        st.subheader("Incentives")
        
        ny_selected = st.radio("Select Incentive Region", ["General", "New York"])
        
        if ny_selected == "New York":
            project_location = st.radio("Project Location", ["Upstate", "Con Edison"])
            
            icsa_eligible = st.checkbox("Eligible for Inclusive Community Solar Adder (ICSA)", 
                help="The ICSA broadens access to community solar for low-to-moderate income households, affordable housing, and disadvantaged communities in New York.")
            if icsa_eligible:
                if project_location == "Upstate":
                    icsa_rate_upstate = st.slider("ICSA Rate (Upstate)", 0.05, 0.20, 0.10, 0.01, 
                        help="For Upstate projects, ICSA rate ranges from $0.05/Watt to $0.20/Watt.")
                else:
                    icsa_rate_con_ed = st.slider("ICSA Rate (Con Edison)", 0.20, 0.30, 0.25, 0.01, 
                        help="For Con Edison projects, ICSA rate ranges from $0.20/Watt to $0.30/Watt.")
            
            brownfield_eligible = st.checkbox("Eligible for Landfill/Brownfield Adder", 
                help="A fixed rate of $0.15/Watt for projects developed on brownfield or landfill sites.")
            
            prevailing_wage_eligible = st.checkbox("Eligible for Prevailing Wage Adder", 
                help="For projects that comply with prevailing wage requirements. Upstate: $0.125/Watt DC, Con Edison: $0.20/Watt DC.")
            
            nyserda_community_eligible = st.checkbox("Eligible for NYSERDA Community Adder", 
                help="A fixed rate of $0.07/Watt for projects in Con Edison and Upstate territories that don't qualify for MTC or CC.")
        
        itc_rate = st.slider('Base Investment Tax Credit (ITC) Rate (%)', min_value=0.0, max_value=30.0, value=30.0, step=0.1, 
            help="Base ITC rate as a percentage of CapEx. Typically 30% for solar projects.") / 100
        
        additional_itc_category = st.radio("Additional ITC Category", 
            ["None", "Category 1 & 2", "Category 3 & 4"], 
            help="""
            Low-Income Communities Bonus Credit Program categories:
            - Category 1: Located in a Low-Income Community (10% additional)
            - Category 2: Located on Indian Land (10% additional)
            - Category 3: Qualified Low-Income Residential Building Project (20% additional)
            - Category 4: Qualified Low-Income Economic Benefit Project (20% additional)
            """)
        if additional_itc_category == "Category 1 & 2":
            additional_itc = 0.10
        elif additional_itc_category == "Category 3 & 4":
            additional_itc = 0.20
        else:
            additional_itc = 0.00
        
        st.write(f"Total ITC Rate: {(itc_rate + additional_itc) * 100:.1f}%")


    incentives = {
        'ny_selected': ny_selected == "New York",
        'project_location': project_location if ny_selected == "New York" else None,
        'icsa_eligible': icsa_eligible if ny_selected == "New York" else False,
        'icsa_rate_upstate': icsa_rate_upstate if ny_selected == "New York" and project_location == "Upstate" and icsa_eligible else 0,
        'icsa_rate_con_ed': icsa_rate_con_ed if ny_selected == "New York" and project_location == "Con Edison" and icsa_eligible else 0,
        'brownfield_eligible': brownfield_eligible if ny_selected == "New York" else False,
        'prevailing_wage_eligible': prevailing_wage_eligible if ny_selected == "New York" else False,
        'nyserda_community_eligible': nyserda_community_eligible if ny_selected == "New York" else False,
        'itc_rate': itc_rate,
        'additional_itc': additional_itc
    }

    project_data = {
    'capacity': project_capacity,
    'capex': capex,
    'opex': opex,
    'lifespan': project_lifespan,
    'degradation_rate': degradation_rate,
    'annual_generation': annual_generation,
    'target_irr': target_irr,
    'inverter_cost': inverter_cost,
    'inverter_lifetime': inverter_lifetime,
    'incentives': incentives,
    'inflation_rate': inflation_rate,
    'incentive_payback_period': incentive_payback_period,
    'use_custom_ppa': use_custom_ppa,
    'custom_ppa_price': custom_ppa_price
    }

    st.session_state.project_data = project_data

    if st.button('Calculate', key='calculate_ppa_button'):
        st.session_state.calculated_results = calculate_results(project_data)

    if st.session_state.calculated_results is not None:
        display_results(st.session_state.calculated_results, project_data)
        
        # Create download button
        excel_buffer = export_to_excel(st.session_state.calculated_results['df'], 
                                       plot_cash_flows(st.session_state.calculated_results['df']))
        st.download_button(
            label="Download Excel Report",
            data=excel_buffer.getvalue(),
            file_name="solar_ppa_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_excel_report"
        )


    st.markdown("""
    ## Mathematical Concepts

    <details>
    <summary>Net Present Value (NPV)</summary>

    The Net Present Value (NPV) is calculated using the following formula:

    NPV = Σ (CF_t / (1+r)^t) for t = 0 to T

    Where:
    - $CF_t$ is the cash flow at time t
    - $r$ is the discount rate (in this case, the target IRR)
    - $T$ is the total number of time periods (project lifespan)

    The goal is to find the PPA price that makes NPV = 0.
    </details>

    <details>
    <summary>Internal Rate of Return (IRR)</summary>

    The Internal Rate of Return (IRR) is the discount rate that makes the NPV of all cash flows equal to zero. It's found by solving the equation:

    0 = Σ (CF_t / (1+IRR)^t) for t = 0 to T

    In our calculator, we use the target IRR as the discount rate and solve for the PPA price that achieves this IRR.
    </details>

    <details>
    <summary>Annual Cash Flow Calculation</summary>

    For each year t, the cash flow is calculated as:

    CF_t = (G_0 * (1-d)^t * P) - OpEx - I_t

    Where:
    - $G_0$ is the initial annual generation
    - $d$ is the annual degradation rate
    - $P$ is the PPA price
    - $OpEx$ is the annual operational expenditure
    - $I_t$ is the inverter replacement cost (if applicable in year t)
    </details>

    <details>
    <summary>Incentive Application</summary>

    Total incentives are calculated as:

    Total Incentives = (CapEx * ITC_rate) + (Capacity * 1000 * Σ(Incentive_rates)) + Upfront_Incentive

    Where:
    - ITC_rate is the Investment Tax Credit rate
    - Incentive_rates include NY-Sun, LMI, Brownfield, Prevailing Wage, NYSERDA Community, and Other rates
    - Upfront_Incentive is any additional lump sum incentive

    These incentives are applied to reduce the effective CapEx in the cash flow calculations.
    </details>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
